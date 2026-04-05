import os
import copy
import datetime
import calendar
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import range_boundaries
from supabase import create_client, Client

# --- SUPABASE CLOUD DATABASE SETUP ---
SUPABASE_URL = "https://uphranefrtukvggisgea.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InVwaHJhbmVmcnR1a3ZnZ2lzZ2VhIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzUzNzY4OTYsImV4cCI6MjA5MDk1Mjg5Nn0.CGl-Q1qFcFB3JCQ-76KJTIp-J1dLIeTbQu9c6iOGit8"

supabase: Client = None
if SUPABASE_URL != "https://uphranefrtukvggisgea.supabase.co":
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# --- FILE PATHS ---
DATA_DIR = "data"
os.makedirs(DATA_DIR, exist_ok=True)

USERS_FILE = os.path.join(DATA_DIR, "users.xlsx")
BOOKINGS_FILE = os.path.join(DATA_DIR, "bookings.xlsx")
CR_TIMETABLE_FILE = os.path.join(DATA_DIR, "cr_timetable.xlsx")
LAB_MERGED_TEMP_FILE = os.path.join(DATA_DIR, "lab_merged_temp.xlsx")
LAB_TIMETABLE_FILE = os.path.join(DATA_DIR, "lab_timetable.xlsx")

# ==========================================
# CLOUD SYNC FUNCTIONS (Prevents data deletion on Render)
# ==========================================
def sync_down(filename):
    """Downloads the file from the cloud when the server wakes up."""
    if not supabase: return
    try:
        data = supabase.storage.from_("smart-class-data").download(filename)
        with open(os.path.join(DATA_DIR, filename), 'wb') as f:
            f.write(data)
    except Exception:
        pass # File doesn't exist on cloud yet, which is fine

def sync_up(filename):
    """Uploads the file to the cloud to keep it safe forever."""
    if not supabase: return
    file_path = os.path.join(DATA_DIR, filename)
    if os.path.exists(file_path):
        try:
            supabase.storage.from_("smart-class-data").remove([filename]) # Delete old version
        except:
            pass
        with open(file_path, 'rb') as f:
            supabase.storage.from_("smart-class-data").upload(path=filename, file=f.read())

# Sync files immediately when the server boots up
sync_down("users.xlsx")
sync_down("bookings.xlsx")
sync_down("cr_timetable.xlsx")
sync_down("lab_timetable.xlsx")

# Ensure persistent logging files exist
if not os.path.exists(USERS_FILE):
    pd.DataFrame(columns=["Email", "Password"]).to_excel(USERS_FILE, index=False)
    sync_up("users.xlsx")

if not os.path.exists(BOOKINGS_FILE):
    pd.DataFrame(columns=["Email", "Type", "Resource", "Date", "Time Slot", "Purpose"]).to_excel(BOOKINGS_FILE, index=False)
    sync_up("bookings.xlsx")

# ==========================================
# 1. AUTHENTICATION & LOGGING
# ==========================================

def register_user(email, password):
    if not email.endswith("@adaniuni.ac.in"):
        return {"error": "Only @adaniuni.ac.in emails are allowed."}
    
    if os.path.exists(USERS_FILE) and os.path.getsize(USERS_FILE) > 0:
        df = pd.read_excel(USERS_FILE)
    else:
        df = pd.DataFrame(columns=["Email", "Password"])
        
    if email in df["Email"].values:
        return {"error": "User already exists."}
        
    new_user = pd.DataFrame([{"Email": email, "Password": password}])
    df = pd.concat([df, new_user], ignore_index=True)
    df.to_excel(USERS_FILE, index=False)
    
    sync_up("users.xlsx") # Save to cloud
    return {"success": "Registered successfully."}

def log_booking(email, resource_type, resource_name, time_slot, date, purpose):
    if os.path.exists(BOOKINGS_FILE) and os.path.getsize(BOOKINGS_FILE) > 0:
        df = pd.read_excel(BOOKINGS_FILE)
    else:
        df = pd.DataFrame(columns=["Email", "Type", "Resource", "Date", "Time Slot", "Purpose"])
        
    new_booking = pd.DataFrame([{
        "Email": email, "Type": resource_type, "Resource": resource_name, 
        "Date": date, "Time Slot": time_slot, "Purpose": purpose
    }])
    df = pd.concat([df, new_booking], ignore_index=True)
    df.to_excel(BOOKINGS_FILE, index=False)
    
    sync_up("bookings.xlsx") # Save to cloud
    return True

# ==========================================
# 2. HELPER UTILITIES
# ==========================================

def time_to_minutes(t):
    try:
        h, m = map(int, str(t).split(":"))
        return h * 60 + m
    except:
        return None

def parse_slot(slot):
    if not slot:
        return None, None
    slot = str(slot).replace("*", "").replace(" ", "").strip()
    if "to" not in slot:
        return None, None
    parts = slot.split("to")
    if len(parts) != 2:
        return None, None
    try:
        return time_to_minutes(parts[0].strip()), time_to_minutes(parts[1].strip())
    except:
        return None, None

def get_real_cell_value(sheet, row, col):
    cell = sheet.cell(row, col)
    value = cell.value
    for merged in sheet.merged_cells.ranges:
        if cell.coordinate in merged:
            value = sheet.cell(merged.min_row, merged.min_col).value
            break
    return value

# ==========================================
# 3. CLASSROOM TIMETABLE GENERATION
# ==========================================

def generate_classroom_full_year(original_file_path, sheet_to_copy, year=2026):
    wb = openpyxl.load_workbook(original_file_path)
    if sheet_to_copy not in wb.sheetnames:
        return False, f"Sheet '{sheet_to_copy}' not found."
        
    source_ws = wb[sheet_to_copy]
    months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    month_num_dict = {m: i+1 for i, m in enumerate(months)}

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for name in months:
        new_ws = wb.copy_worksheet(source_ws)
        new_ws.title = name
        new_ws._images = []
        for img in source_ws._images:
            new_ws.add_image(copy.deepcopy(img))

    for month in months:
        ws = wb[month]
        month_num = month_num_dict[month]
        days_in_month = calendar.monthrange(year, month_num)[1]
        first_day = datetime.date(year, month_num, 1)
        first_weekday = first_day.weekday()
        mon_date = 1 - first_weekday

        header_rows = []
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 1).value and "Classroom No :" in str(ws.cell(row, 1).value):
                header_rows.append(row)

        for hr in header_rows:
            merge_range = f"B{hr}:F{hr}"
            if any(mrg.coord == merge_range for mrg in ws.merged_cells.ranges):
                ws.unmerge_cells(merge_range)

        ws.insert_cols(7, amount=1)

        for r in range(1, ws.max_row + 1):
            fri_cell = ws.cell(row=r, column=6)
            sat_cell = ws.cell(row=r, column=7)
            
            if fri_cell.has_style:
                sat_cell.font = copy.copy(fri_cell.font)
                sat_cell.border = copy.copy(fri_cell.border)
                sat_cell.fill = copy.copy(fri_cell.fill)
                sat_cell.number_format = fri_cell.number_format
                sat_cell.protection = copy.copy(fri_cell.protection)
                sat_cell.alignment = copy.copy(fri_cell.alignment)

        merges_to_expand = []
        for mrg in list(ws.merged_cells.ranges):
            if mrg.min_col == 2 and mrg.max_col == 6 and mrg.min_row not in header_rows:
                merges_to_expand.append(mrg)

        for mrg in merges_to_expand:
            ws.unmerge_cells(mrg.coord)
            ws.merge_cells(start_row=mrg.min_row, start_column=2, end_row=mrg.max_row, end_column=7)

        for r in [1, 2, 3]:
            for mrg in list(ws.merged_cells.ranges):
                if mrg.min_row == r and mrg.max_row == r:
                    ws.unmerge_cells(mrg.coord)
            
            title_text = ws.cell(row=r, column=1).value
            if not title_text: 
                title_text = ws.cell(row=r, column=2).value
            
            ws.cell(row=r, column=1).value = title_text
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')

        for hr in header_rows:
            ws.cell(row=hr+1, column=7).value = "SAT"
            ws.cell(row=hr+1, column=7).font = Font(bold=True)
            ws.cell(row=hr+1, column=7).alignment = Alignment(horizontal='center')

            for col_offset in range(6): 
                day_num = mon_date + col_offset
                cell = ws.cell(row=hr, column=2 + col_offset)
                if 1 <= day_num <= days_in_month:
                    cell.value = day_num
                    cell.font = Font(bold=True, color="000000")
                else:
                    cell.value = None
                    cell.font = Font(bold=True, color="AAAAAA")
                cell.alignment = Alignment(horizontal='center')

        merged_rows = set()
        for mrg in merges_to_expand:
            for r in range(mrg.min_row, mrg.max_row + 1):
                merged_rows.add(r)

        for r in range(1, ws.max_row + 1):
            if r <= 3: continue  
            if r in header_rows: continue  
            if r in [hr + 1 for hr in header_rows]: continue  
            if r in merged_rows: continue  
            
            cell = ws.cell(row=r, column=7) 
            cell.value = "NA"
            cell.fill = yellow_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        fri_col_letter = openpyxl.utils.get_column_letter(6)
        if fri_col_letter in ws.column_dimensions:
            fri_width = ws.column_dimensions[fri_col_letter].width
            ws.column_dimensions[openpyxl.utils.get_column_letter(7)].width = fri_width

        group_mon_date = mon_date + 7
        while True:
            if not any(1 <= group_mon_date + off <= days_in_month for off in range(6)):
                break

            insert_col = ws.max_column + 1
            ws.insert_cols(insert_col, amount=6)

            for r in range(1, ws.max_row + 1):
                for off in range(6):
                    src = ws.cell(r, 2 + off)
                    tgt = ws.cell(r, insert_col + off)
                    tgt.value = src.value
                    if src.has_style:
                        tgt.font = copy.copy(src.font)
                        tgt.border = copy.copy(src.border)
                        tgt.fill = copy.copy(src.fill)
                        tgt.number_format = src.number_format
                        tgt.protection = copy.copy(src.protection)
                        tgt.alignment = copy.copy(src.alignment)

            for off in range(6):
                src_col = openpyxl.utils.get_column_letter(2 + off)
                tgt_col = openpyxl.utils.get_column_letter(insert_col + off)
                if src_col in ws.column_dimensions:
                    ws.column_dimensions[tgt_col].width = ws.column_dimensions[src_col].width

            offset = insert_col - 2
            for mrg in list(ws.merged_cells.ranges):
                if 2 <= mrg.min_col <= mrg.max_col <= 7:
                    ws.merge_cells(
                        start_row=mrg.min_row, start_column=mrg.min_col + offset,
                        end_row=mrg.max_row, end_column=mrg.max_col + offset
                    )

            for r in [1, 2, 3]:
                ws.merge_cells(start_row=r, start_column=insert_col, end_row=r, end_column=insert_col + 5)
                tgt_cell = ws.cell(row=r, column=insert_col)
                src_cell = ws.cell(row=r, column=1) 
                
                tgt_cell.value = src_cell.value
                if src_cell.has_style:
                    tgt_cell.font = copy.copy(src_cell.font)
                    tgt_cell.fill = copy.copy(src_cell.fill)
                    tgt_cell.border = copy.copy(src_cell.border)
                tgt_cell.alignment = Alignment(horizontal='center', vertical='center')

            for hr in header_rows:
                for col_offset in range(6):
                    day_num = group_mon_date + col_offset
                    cell = ws.cell(row=hr, column=insert_col + col_offset)
                    if 1 <= day_num <= days_in_month:
                        cell.value = day_num
                        cell.font = Font(bold=True, color="000000")
                    else:
                        cell.value = None
                        cell.font = Font(bold=True, color="AAAAAA")
                    cell.alignment = Alignment(horizontal='center')

            group_mon_date += 7

    wb.save(CR_TIMETABLE_FILE)
    sync_up("cr_timetable.xlsx") # Save to cloud
    return True, "Classroom full-year timetable generated."

# ==========================================
# 4. LAB TIMETABLE GENERATION
# ==========================================

def generate_vertically_merged_lab(original_file_path):
    wb = openpyxl.load_workbook(original_file_path)
    sheet_names = wb.sheetnames
    
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "Merged_Sheet"
    
    current_row = 1
    
    for i, sheet_name in enumerate(sheet_names):
        ws = wb[sheet_name]
        start_row = 1 if i == 0 else 2
        
        last_row = 0
        for row in ws.iter_rows():
            if any(cell.value is not None for cell in row):
                last_row = row[0].row
                
        if last_row < start_row:
            continue
            
        row_offset = current_row - start_row
        
        source_merged_ranges = []
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if max_row >= start_row and min_row <= last_row:
                source_merged_ranges.append(merged_range)
                
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=last_row)):
            current_source_row = row_idx + start_row
            for cell in row:
                target_row = current_source_row + row_offset
                target_column = cell.column
                
                is_part_of_merged_range = False
                is_top_left_of_merged_range = False
                for m_range in source_merged_ranges:
                    m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(m_range))
                    if m_min_row <= current_source_row <= m_max_row and m_min_col <= cell.column <= m_max_col:
                        is_part_of_merged_range = True
                        if current_source_row == m_min_row and cell.column == m_min_col:
                            is_top_left_of_merged_range = True
                        break
                
                new_cell = new_ws.cell(row=target_row, column=target_column)
                if not is_part_of_merged_range or is_top_left_of_merged_range:
                    new_cell.value = cell.value
                
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = copy.copy(cell.number_format)
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)
                    
        for merged_range in source_merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            new_min_row = min_row + row_offset
            new_max_row = max_row + row_offset
            new_range = openpyxl.utils.get_column_letter(min_col) + str(new_min_row) + ":" + \
                        openpyxl.utils.get_column_letter(max_col) + str(new_max_row)
            new_ws.merge_cells(new_range)
            
        current_row += last_row - start_row + 1
        
    new_wb.save(LAB_MERGED_TEMP_FILE)
    return LAB_MERGED_TEMP_FILE


def generate_lab_full_year(merged_file_path, year=2026):
    wb = openpyxl.load_workbook(merged_file_path)
    sheet_to_copy = "Merged_Sheet"
    
    if sheet_to_copy not in wb.sheetnames:
        return False, "Merged_Sheet not found in intermediate file."
        
    source_ws = wb[sheet_to_copy]
    months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    month_num_dict = {m: i+1 for i, m in enumerate(months)}

    for name in months:
        new_ws = wb.copy_worksheet(source_ws)
        new_ws.title = name

    for month in months:
        ws = wb[month]
        month_num = month_num_dict[month]
        days_in_month = calendar.monthrange(year, month_num)[1]

        first_day = datetime.date(year, month_num, 1)
        first_weekday = first_day.weekday()
        mon_date = 1 - first_weekday

        header_rows = []
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r,2).value
            if val and str(val).strip().upper() == "MON":
                header_rows.append(r)

        horizontal_merges = []
        for mrg in list(ws.merged_cells.ranges):
            if mrg.min_row == mrg.max_row and 1 <= mrg.min_col <= 6:
                horizontal_merges.append((mrg.min_row, mrg.min_col, mrg.max_col))

        ws.insert_cols(7, amount=1)

        for r in range(1, ws.max_row + 1):
            fri_cell = ws.cell(row=r, column=6)
            sat_cell = ws.cell(row=r, column=7)
            
            if fri_cell.has_style:
                sat_cell.font = copy.copy(fri_cell.font)
                sat_cell.border = copy.copy(fri_cell.border)
                sat_cell.fill = copy.copy(fri_cell.fill)
                sat_cell.number_format = fri_cell.number_format
                sat_cell.protection = copy.copy(fri_cell.protection)
                sat_cell.alignment = copy.copy(fri_cell.alignment)

        fri_col_letter = openpyxl.utils.get_column_letter(6)
        if fri_col_letter in ws.column_dimensions:
            fri_width = ws.column_dimensions[fri_col_letter].width
            ws.column_dimensions[openpyxl.utils.get_column_letter(7)].width = fri_width

        friday_vertical_merges = []
        for mrg in list(ws.merged_cells.ranges):
            if mrg.min_col == 6 and mrg.max_col == 6 and mrg.min_row != mrg.max_row:
                friday_vertical_merges.append((mrg.min_row, mrg.max_row))

        for min_r, max_r in friday_vertical_merges:
            ws.merge_cells(start_row=min_r, start_column=7, end_row=max_r, end_column=7) 

        for hr in header_rows:
            for col_offset in range(6): 
                day_num = mon_date + col_offset
                cell = ws.cell(row=hr, column=2 + col_offset)
                if 1 <= day_num <= days_in_month:
                    cell.value = day_num
                    cell.font = Font(bold=True)
                else:
                    cell.value = None
                cell.alignment = Alignment(horizontal='center')

        group_mon_date = mon_date + 7

        while True:
            if not any(1 <= group_mon_date + off <= days_in_month for off in range(6)):
                break

            insert_col = ws.max_column + 1
            ws.insert_cols(insert_col, amount=6) 

            for r in range(1, ws.max_row + 1):
                for off in range(6):
                    src = ws.cell(r, 2 + off)
                    tgt = ws.cell(r, insert_col + off)
                    tgt.value = src.value

                    if src.has_style:
                        tgt.font = copy.copy(src.font)
                        tgt.border = copy.copy(src.border)
                        tgt.fill = copy.copy(src.fill)
                        tgt.number_format = src.number_format
                        tgt.protection = copy.copy(src.protection)
                        tgt.alignment = copy.copy(src.alignment)

            for off in range(6): 
                src_col = openpyxl.utils.get_column_letter(2 + off)
                tgt_col = openpyxl.utils.get_column_letter(insert_col + off)
                if src_col in ws.column_dimensions:
                    ws.column_dimensions[tgt_col].width = ws.column_dimensions[src_col].width

            offset = insert_col - 2
            for mrg in list(ws.merged_cells.ranges):
                if 2 <= mrg.min_col <= mrg.max_col <= 7 and mrg.min_row != mrg.max_row: 
                    ws.merge_cells(
                        start_row=mrg.min_row,
                        start_column=mrg.min_col + offset,
                        end_row=mrg.max_row,
                        end_column=mrg.max_col + offset
                    )

            for hr in header_rows:
                for col_offset in range(6):
                    day_num = group_mon_date + col_offset
                    cell = ws.cell(row=hr, column=insert_col + col_offset)
                    if 1 <= day_num <= days_in_month:
                        cell.value = day_num
                        cell.font = Font(bold=True)
                    else:
                        cell.value = None
                    cell.alignment = Alignment(horizontal='center')

            group_mon_date += 7

        last_col = ws.max_column

        for row, col, end_col in horizontal_merges:
            try:
                mrg_to_remove = None
                for mrg in list(ws.merged_cells.ranges):
                    if mrg.min_row == row and mrg.min_col == col:
                        mrg_to_remove = mrg
                        break
                if mrg_to_remove:
                    ws.unmerge_cells(mrg_to_remove.coord)
            except:
                pass

            ws.merge_cells(
                start_row=row,
                start_column=col,
                end_row=row,
                end_column=last_col
            )

    wb.save(LAB_TIMETABLE_FILE)
    sync_up("lab_timetable.xlsx") # Save to cloud
    
    if os.path.exists(LAB_MERGED_TEMP_FILE):
        os.remove(LAB_MERGED_TEMP_FILE)
        
    return True, "Lab full-year timetable generated."

# ==========================================
# 5. SEARCH & BOOKING LOGIC
# ==========================================

def search_free_slots(date_str, start_time_str, end_time_str, resource_type):
    file_to_check = CR_TIMETABLE_FILE if resource_type == "Classroom" else LAB_TIMETABLE_FILE
    
    if not os.path.exists(file_to_check):
        return {"error": f"{resource_type} timetable not uploaded yet."}

    selected_date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
    month_name = selected_date.strftime("%b")
    day_number = selected_date.day
    weekday_name = selected_date.strftime("%a").upper()

    wb = openpyxl.load_workbook(file_to_check, data_only=True)
    if month_name not in wb.sheetnames:
        return {"error": f"Sheet '{month_name}' not found."}
    
    sheet = wb[month_name]
    target_column = None

    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value).strip() == str(day_number):
                if resource_type == "Classroom":
                    below = sheet.cell(cell.row + 1, cell.column).value
                    if below and weekday_name in str(below).upper():
                        target_column = cell.column
                        break
                else:
                    target_column = cell.column
                    break
        if target_column:
            break

    if not target_column:
        return {"error": "Could not find column for the selected date."}

    user_start = time_to_minutes(start_time_str)
    user_end = time_to_minutes(end_time_str)
    if user_start == 0 and user_end == 0:
        user_end = 23 * 60 + 59

    free_slots = []
    rows = list(sheet.iter_rows())
    i = 0

    while i < len(rows):
        first_cell = rows[i][0].value
        resource_indicator = "Classroom No" if resource_type == "Classroom" else "Lab"
        
        if first_cell and resource_indicator in str(first_cell):
            resource_full = str(first_cell).split(":", 1)[-1].strip() if resource_type == "Classroom" else str(first_cell).strip()
            j = rows[i][0].row + 2

            while j <= sheet.max_row:
                time_cell = sheet.cell(j, 1).value
                if time_cell and resource_indicator in str(time_cell):
                    break
                if resource_type == "Lab" and time_cell and ("Break" in str(time_cell) or "Recess" in str(time_cell)):
                    j += 1
                    continue

                slot_start, slot_end = parse_slot(time_cell)
                if slot_start is None or (slot_end - slot_start) < 40:
                    j += 1
                    continue

                if slot_start < user_end and slot_end > user_start:
                    cell_value = get_real_cell_value(sheet, j, target_column) if resource_type == "Lab" else sheet.cell(j, target_column).value
                    status = str(cell_value).strip().upper() if cell_value else ""

                    if status in ("", "NA", "NONE"):
                        free_slots.append({
                            "resource": resource_full, 
                            "row": j, 
                            "time_slot": str(time_cell).strip(),
                            "month": month_name,
                            "target_column": target_column
                        })
                j += 1
            i = j - 2
        i += 1

    return {"slots": free_slots}

def book_slots_in_excel(slots_to_book, month_name, target_column, purpose, resource_type):
    file_to_check = CR_TIMETABLE_FILE if resource_type == "Classroom" else LAB_TIMETABLE_FILE
    
    wb = openpyxl.load_workbook(file_to_check)
    sheet = wb[month_name]
    
    for row_num in slots_to_book:
        sheet.cell(row=row_num, column=target_column).value = purpose
        
    wb.save(file_to_check)
    
    # Save the updated timetable back to the cloud
    sync_up("cr_timetable.xlsx" if resource_type == "Classroom" else "lab_timetable.xlsx")
    return True

# ==========================================
# 6. UTILITY ANALYSIS LOGIC
# ==========================================

def get_resource_names(resource_type):
    file_to_check = CR_TIMETABLE_FILE if resource_type == "Classroom" else LAB_TIMETABLE_FILE
    
    if not os.path.exists(file_to_check):
        return []

    wb = openpyxl.load_workbook(file_to_check, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    
    names = []
    for row in sheet.iter_rows(min_col=1, max_col=1):
        val = row[0].value
        if val:
            text = str(val)
            if resource_type == "Classroom" and "Classroom No" in text:
                names.append(text.split(":")[-1].strip())
            elif resource_type == "Lab" and "Lab" in text:
                names.append(text.strip())
                
    return list(dict.fromkeys(names))


def calculate_utility(resource_type, selected_resource):
    file_to_check = CR_TIMETABLE_FILE if resource_type == "Classroom" else LAB_TIMETABLE_FILE
    
    if not os.path.exists(file_to_check):
        return {"error": f"{resource_type} timetable not uploaded yet."}

    wb = openpyxl.load_workbook(file_to_check, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    
    occupied, free = 0, 0
    rows = list(sheet.iter_rows())
    merged_ranges = sheet.merged_cells.ranges
    i = 0

    while i < len(rows):
        first_cell = rows[i][0].value
        if first_cell:
            text = str(first_cell)
            match = False

            if resource_type == "Classroom" and "Classroom No" in text:
                name = text.split(":")[-1].strip()
                if selected_resource == "All" or name == selected_resource:
                    match = True
            elif resource_type == "Lab" and "Lab" in text:
                name = text.strip()
                if selected_resource == "All" or name == selected_resource:
                    match = True

            if match:
                j = rows[i][0].row + 2
                while j <= sheet.max_row:
                    time_val = sheet.cell(j, 1).value
                    if time_val and ("Classroom No" in str(time_val) or "Lab" in str(time_val)):
                        break
                    
                    is_break = False
                    for col in range(1, 8): 
                        val = sheet.cell(j,col).value
                        if val and ("break" in str(val).lower() or "recess" in str(val).lower()):
                            is_break = True
                            break
                            
                    if is_break or "to" not in str(time_val):
                        j += 1
                        continue

                    slot_height = 1
                    for m in merged_ranges:
                        if m.min_row == j and m.min_col == 2:
                            slot_height = m.max_row - m.min_row + 1
                            break

                    for col in range(2, 8): 
                        occupied_flag = False
                        for r in range(j, j + slot_height):
                            val = sheet.cell(r, col).value
                            if val and str(val).strip().upper() not in ["", "NA"]:
                                occupied_flag = True
                                break
                        if occupied_flag:
                            occupied += 1
                        else:
                            free += 1
                    j += slot_height
                i = j - 2
        i += 1

    return {"occupied": occupied, "free": free, "total": occupied + free}